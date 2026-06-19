"""
Studio Ritardi — BUILD file Excel WC (league_id=1) COMPATIBILE OFFICE 2013.

Filosofia (il FILE e' l'oracolo, non la dashboard):
- DATI MATCH popolato con TUTTI gli eventi WC grezzi (input: risultati 90' + PT).
- Tutte le formule CLASSICHE del foglio restano VIVE: il tuo Excel 2013 ricalcola
  da solo W/L, RIT, SUC, le StatCard (Quota/Ritardo/%/Media Rit), distribuzione
  F/G/H e sotto/sopra media. Nessun numero calcolato da noi in quelle celle.
- Le UNICHE 10 celle per foglio incompatibili 2013 (FILTER/UNIQUE/SEQUENCE) sono
  neutralizzate; i 3 blocchi-presentazione che dipendono da esse (ultime-10,
  storico-serie per frequenza, run-sopra-media) sono riempiti con i VALORI gia'
  certificati 1:1 vs dashboard (_cert_wc_compare = 50/50). Calcolati sul target
  di default del foglio; se cambi il target, le StatCard si aggiornano dal vivo,
  questi 3 blocchi restano lo snapshot del default.
"""
import json
import os
import sys

import openpyxl
from openpyxl.formula.translate import Translator
from openpyxl.utils import get_column_letter, column_index_from_string, range_boundaries
from openpyxl.worksheet.formula import ArrayFormula

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from _cert_wc_compare import oracle  # type: ignore  # noqa

BASE = r"C:\Users\Admin\Desktop\PYTHON DATABASE"
SRC = os.path.join(BASE, "STUDIO RITARDI_BASE_v5.0 - FORMULE LIBERE.xlsx")
OUT = os.path.join(BASE, "STUDIO RITARDI_CERT_WC_league1.xlsx")
DATA = os.path.join(os.path.dirname(os.path.abspath(__file__)), "_cert_wc_data.json")

HEADER_ROW = 4
DATA_START = 5
# colonne dati grezzi (lettera) nel foglio DATI MATCH
COL = dict(evento="B", home="C", away="D", gc="H", ga="I", gcfh="J", gafh="K")

# foglio -> (market, target_default, [(input_cell, value), ...])
SHEETS = {
    "RIS.ESATTI": ("re", "1-1", [("C9", 1), ("D9", 1)]),
    "SGE":        ("sge", "3", [("C8", 3)]),
    "OVER":       ("over", "2.5", [("C8", 2.5)]),
    "UNDER":      ("under", "2.5", [("C8", 2.5)]),
    "GGPT":       ("ggpt", None, []),
    "GGST":       ("ggst", None, []),
    "OVPT":       ("ovpt", "1.5", [("C8", 1.5)]),
    "PF1X":       ("pf1x", None, []),
    "PF2X":       ("pf2x", None, []),
    "PFX1":       ("pfx1", None, []),
    "PFX2":       ("pfx2", None, []),
    "X":          ("x", None, []),
    "GG&OV25":    ("ggov25", None, []),
}

# 10 celle dinamiche (FILTER/UNIQUE/SEQUENCE) presenti, identiche, in ogni foglio
DYN_CELLS = ["BB3", "BC3", "BD3", "BE3", "BF3", "BG3", "BN3", "AZ4", "BL4", "BR4"]


def fval(cell):
    v = cell.value
    return v.text if isinstance(v, ArrayFormula) else v


def is_formula(cell):
    v = cell.value
    if isinstance(v, ArrayFormula):
        return True
    return isinstance(v, str) and v.startswith("=")


def populate_dati_match(ws, events):
    n = len(events)
    last_row = DATA_START + n - 1
    min_c, _, max_c, _ = range_boundaries(ws.tables["Tabella1"].ref)
    ci = {k: column_index_from_string(v) for k, v in COL.items()}
    data_cols = set(ci.values()) | {column_index_from_string(x) for x in ("E", "F", "G")}

    # colonne-formula da propagare (template = riga 6, ricorsiva) escluse le data-cols
    formula_cols = [c for c in range(min_c, max_c + 1)
                    if c not in data_cols and is_formula(ws.cell(DATA_START + 1, c))]

    # 1) dati grezzi
    for i, e in enumerate(events):
        r = DATA_START + i
        ws.cell(r, ci["evento"]).value = i + 1
        ws.cell(r, ci["home"]).value = e["home"]
        ws.cell(r, ci["away"]).value = e["away"]
        ws.cell(r, ci["gc"]).value = e["gc"]
        ws.cell(r, ci["ga"]).value = e["ga"]
        ws.cell(r, ci["gcfh"]).value = e["gcfh"]
        ws.cell(r, ci["gafh"]).value = e["gafh"]

    # 2) propaga formule classiche per ogni riga dati >=6 (riga 5 = template base, lasciata)
    for c in formula_cols:
        letter = get_column_letter(c)
        tmpl = fval(ws.cell(DATA_START + 1, c))      # riga 6
        origin = f"{letter}{DATA_START + 1}"
        for r in range(DATA_START + 1, last_row + 1):
            ws.cell(r, c).value = Translator(tmpl, origin=origin).translate_formula(f"{letter}{r}")
    # svuota eventuali righe-formula residue oltre i dati (se il template ne aveva)
    for c in range(min_c, max_c + 1):
        for r in range(last_row + 1, ws.max_row + 1):
            if is_formula(ws.cell(r, c)):
                ws.cell(r, c).value = None

    # 3) estendi Tabella1
    ws.tables["Tabella1"].ref = f"{get_column_letter(min_c)}{HEADER_ROW}:{get_column_letter(max_c)}{last_row}"
    return last_row


def fill_blocks(ws, orc):
    """Riempie i 3 blocchi (valori 1:1 dashboard) e neutralizza le celle dinamiche."""
    # neutralizza le 10 celle dinamiche -> i loro dipendenti classici degradano a "" (no #NOME?)
    for coord in DYN_CELLS:
        ws[coord].value = None

    # --- ULTIME 10 SERIE: AZ4:AZ13 + BL4:BL13 (oldest..newest) ---
    u = orc["ultime10"]
    for j in range(10):
        v = u[j] if j < len(u) else None
        ws.cell(4 + j, column_index_from_string("AZ")).value = v
        ws.cell(4 + j, column_index_from_string("BL")).value = v

    # --- STORICO SERIE (per frequenza desc, tie valore asc): BE/BF/BG da riga 3 ---
    be, bf, bg = (column_index_from_string(x) for x in ("BE", "BF", "BG"))
    for i, s in enumerate(orc["storico"]):
        r = 3 + i
        ws.cell(r, be).value = s["len"]
        ws.cell(r, bf).value = s["count"]
        ws.cell(r, bg).value = round(s["pct"], 6) if s["pct"] is not None else None

    # --- RUN SOPRA MEDIA: BT/BU/BV da riga 4 ---
    bt, bu, bv = (column_index_from_string(x) for x in ("BT", "BU", "BV"))
    for i, rr in enumerate(orc["run_hist"]):
        r = 4 + i
        ws.cell(r, bt).value = rr["run_len"]
        ws.cell(r, bu).value = rr["count"]
        ws.cell(r, bv).value = round(rr["pct"], 6) if rr["pct"] is not None else None


def main():
    events = json.load(open(DATA, encoding="utf-8"))
    print(f"eventi WC: {len(events)}  ({events[0]['fixture_date'][:10]} -> {events[-1]['fixture_date'][:10]})")

    wb = openpyxl.load_workbook(SRC, data_only=False)

    last_row = populate_dati_match(wb["DATI MATCH"], events)
    print(f"DATI MATCH popolato: righe {DATA_START}..{last_row}, Tabella1={wb['DATI MATCH'].tables['Tabella1'].ref}")

    for sheet, (market, target, inputs) in SHEETS.items():
        ws = wb[sheet]
        for cell, val in inputs:
            ws[cell].value = val
        orc = oracle(events, market, target)
        fill_blocks(ws, orc)
        tlabel = f" {target}" if target else ""
        print(f"  {sheet:11s} {market}{tlabel:6s} -> n_occ={orc['n_occ']:>3} q.ogg="
              f"{(orc['media_storica'] or 0):.2f} rit={orc['ritardo_attuale']:>3} rec={orc['record']} "
              f"storico={len(orc['storico'])} run={len(orc['run_hist'])} ult10={len(orc['ultime10'])}")

    wb.calculation.fullCalcOnLoad = True
    wb.calculation.calcMode = "auto"
    wb.save(OUT)
    print(f"\nSALVATO -> {OUT}")


if __name__ == "__main__":
    main()
